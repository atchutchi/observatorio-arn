"""
Serviço de Exportação de Relatórios ARN
Suporta exportação em PDF, Excel e CSV
"""
from io import BytesIO
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.template.loader import get_template
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import csv


class ARNExportService:
    """Serviço centralizado para exportação de relatórios"""
    
    # Cores do tema ARN/Binance
    COLORS = {
        'primary': colors.HexColor('#F0B90B'),
        'secondary': colors.HexColor('#181A20'),
        'text': colors.HexColor('#EAECEF'),
        'header': colors.HexColor('#2B3139'),
        'border': colors.HexColor('#474D57'),
    }
    
    def __init__(self, report_data, year, report_type='market'):
        self.report_data = report_data
        self.year = year
        self.report_type = report_type
        self.timestamp = datetime.now()
    
    def export_to_pdf(self):
        """Exporta relatório para PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )
        
        # Container para elementos do PDF
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado para título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=self.COLORS['primary'],
            spaceAfter=30,
            alignment=TA_CENTER,
        )
        
        # Estilo para subtítulos
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.black,
            spaceAfter=12,
        )
        
        # Título do Relatório
        title_text = self._get_report_title()
        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 12))
        
        # Informações gerais
        info_text = f"Ano: {self.year} | Gerado em: {self.timestamp.strftime('%d/%m/%Y às %H:%M')}"
        elements.append(Paragraph(info_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Adicionar conteúdo específico por tipo de relatório
        if self.report_type == 'market':
            elements.extend(self._build_market_report_pdf(styles, subtitle_style))
        elif self.report_type == 'executive':
            elements.extend(self._build_executive_report_pdf(styles, subtitle_style))
        elif self.report_type == 'comparative':
            elements.extend(self._build_comparative_report_pdf(styles, subtitle_style))
        
        # Rodapé
        elements.append(Spacer(1, 30))
        footer_text = "Autoridade Reguladora Nacional (ARN) - Guiné-Bissau"
        elements.append(Paragraph(footer_text, styles['Normal']))
        
        # Gerar PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        
        return pdf
    
    def export_to_excel(self):
        """Exporta relatório para Excel"""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")
        
        wb = Workbook()
        
        # Remover sheet padrão
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Criar sheets por seção
        if self.report_type == 'market':
            self._build_market_report_excel(wb)
        elif self.report_type == 'executive':
            self._build_executive_report_excel(wb)
        elif self.report_type == 'comparative':
            self._build_comparative_report_excel(wb)
        
        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        excel = buffer.getvalue()
        buffer.close()
        
        return excel
    
    def export_to_csv(self):
        """Exporta relatório para CSV"""
        buffer = BytesIO()
        writer = csv.writer(buffer)
        
        # Header
        writer.writerow([f'Relatório ARN - {self.year}'])
        writer.writerow([f'Gerado em: {self.timestamp.strftime("%d/%m/%Y %H:%M")}'])
        writer.writerow([])
        
        # Conteúdo específico
        if self.report_type == 'market':
            self._build_market_report_csv(writer)
        elif self.report_type == 'executive':
            self._build_executive_report_csv(writer)
        elif self.report_type == 'comparative':
            self._build_comparative_report_csv(writer)
        
        csv_data = buffer.getvalue()
        buffer.close()
        
        return csv_data
    
    # ===== MÉTODOS AUXILIARES =====
    
    def _get_report_title(self):
        """Retorna título do relatório baseado no tipo"""
        titles = {
            'market': 'Relatório de Mercado de Telecomunicações',
            'executive': 'Dashboard Executivo ARN',
            'comparative': 'Análise Comparativa entre Operadoras',
        }
        return titles.get(self.report_type, 'Relatório ARN')
    
    def _format_number(self, value):
        """Formata números para exibição"""
        if value is None:
            return "0"
        if isinstance(value, (int, float, Decimal)):
            return f"{value:,.0f}".replace(',', '.')
        return str(value)
    
    def _format_currency(self, value):
        """Formata valores monetários"""
        if value is None:
            return "0 FCFA"
        return f"{self._format_number(value)} FCFA"
    
    # ===== CONSTRUÇÃO DE RELATÓRIO DE MERCADO =====
    
    def _build_market_report_pdf(self, styles, subtitle_style):
        """Constrói relatório de mercado em PDF"""
        elements = []
        
        # Panorama Geral
        panorama = self.report_data.get('panorama_geral', {})
        if panorama:
            elements.append(Paragraph("1. Panorama Geral do Mercado", subtitle_style))
            elements.append(Spacer(1, 12))
            
            data = [
                ['Indicador', 'Valor'],
                ['Total de Assinantes', self._format_number(panorama.get('total_assinantes', 0))],
                ['Crescimento Anual', f"{panorama.get('crescimento_percentual', 0):.2f}%"],
                ['Taxa de Penetração', f"{panorama.get('taxa_penetracao', 0):.2f}%"],
            ]
            
            # Market Share
            for operadora, share in panorama.get('market_share', {}).items():
                data.append([f'Market Share {operadora}', f'{share:.2f}%'])
            
            table = Table(data, colWidths=[3*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.COLORS['header']),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, self.COLORS['border'])
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
        
        # Tráfego de Voz
        trafego = self.report_data.get('trafego_voz', {})
        if trafego:
            elements.append(Paragraph("2. Tráfego de Voz", subtitle_style))
            elements.append(Spacer(1, 12))
            
            data = [
                ['Tipo de Tráfego', 'Volume (minutos)', 'Percentual'],
            ]
            
            distribuicao = trafego.get('distribuicao_percentual', {})
            data.append(['On-net', self._format_number(trafego.get('volume_total', 0)), 
                        f"{distribuicao.get('on_net', 0):.1f}%"])
            data.append(['Off-net', '', f"{distribuicao.get('off_net', 0):.1f}%"])
            data.append(['Internacional', '', f"{distribuicao.get('internacional', 0):.1f}%"])
            
            table = Table(data, colWidths=[2*inch, 2*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.COLORS['header']),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, self.COLORS['border'])
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
        
        # Receitas
        receitas = self.report_data.get('receitas', {})
        if receitas:
            elements.append(Paragraph("3. Receitas", subtitle_style))
            elements.append(Spacer(1, 12))
            
            data = [
                ['Operadora', 'Receita Total (FCFA)', 'Market Share'],
            ]
            
            for operadora, valor in receitas.get('por_operadora', {}).items():
                share = receitas.get('quota_receitas', {}).get(operadora, 0)
                data.append([operadora, self._format_number(valor), f'{share:.1f}%'])
            
            table = Table(data, colWidths=[2*inch, 2.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.COLORS['header']),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, self.COLORS['border'])
            ]))
            
            elements.append(table)
        
        return elements
    
    def _build_executive_report_pdf(self, styles, subtitle_style):
        """Constrói dashboard executivo em PDF"""
        elements = []
        
        kpis = self.report_data.get('kpis_principais', {})
        
        elements.append(Paragraph("KPIs Principais", subtitle_style))
        elements.append(Spacer(1, 12))
        
        data = [['Indicador', 'Valor', 'Variação']]
        
        for key, kpi in kpis.items():
            nome = key.replace('_', ' ').title()
            valor = kpi.get('valor', '--')
            variacao = kpi.get('variacao', '--')
            data.append([nome, str(valor), str(variacao)])
        
        table = Table(data, colWidths=[2.5*inch, 2*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.COLORS['header']),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, self.COLORS['border'])
        ]))
        
        elements.append(table)
        
        return elements
    
    def _build_comparative_report_pdf(self, styles, subtitle_style):
        """Constrói relatório comparativo em PDF"""
        elements = []
        
        elements.append(Paragraph("Comparação entre Operadoras", subtitle_style))
        elements.append(Spacer(1, 12))
        
        if isinstance(self.report_data, list):
            data = [['Operadora', 'Assinantes', 'Market Share', 'Tráfego Voz']]
            
            for item in self.report_data:
                data.append([
                    item.get('operadora', '--'),
                    self._format_number(item.get('assinantes', 0)),
                    item.get('market_share', '--'),
                    self._format_number(item.get('trafego_voz', 0))
                ])
            
            table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.COLORS['header']),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, self.COLORS['border'])
            ]))
            
            elements.append(table)
        
        return elements
    
    # ===== CONSTRUÇÃO DE RELATÓRIOS EXCEL =====
    
    def _build_market_report_excel(self, wb):
        """Constrói relatório de mercado em Excel"""
        # Sheet 1: Panorama Geral
        ws1 = wb.create_sheet("Panorama Geral")
        self._style_excel_header(ws1, f"Relatório de Mercado - {self.year}")
        
        panorama = self.report_data.get('panorama_geral', {})
        
        row = 3
        ws1[f'A{row}'] = "Total de Assinantes"
        ws1[f'B{row}'] = panorama.get('total_assinantes', 0)
        
        row += 1
        ws1[f'A{row}'] = "Crescimento Anual"
        ws1[f'B{row}'] = f"{panorama.get('crescimento_percentual', 0):.2f}%"
        
        row += 1
        ws1[f'A{row}'] = "Taxa de Penetração"
        ws1[f'B{row}'] = f"{panorama.get('taxa_penetracao', 0):.2f}%"
        
        # Market Share
        row += 2
        ws1[f'A{row}'] = "Market Share"
        ws1[f'A{row}'].font = Font(bold=True)
        
        for operadora, share in panorama.get('market_share', {}).items():
            row += 1
            ws1[f'A{row}'] = operadora
            ws1[f'B{row}'] = f'{share:.2f}%'
        
        self._auto_size_columns(ws1)
        
        # Sheet 2: Receitas
        ws2 = wb.create_sheet("Receitas")
        self._style_excel_header(ws2, "Receitas por Operadora")
        
        receitas = self.report_data.get('receitas', {})
        
        ws2['A3'] = "Operadora"
        ws2['B3'] = "Receita Total (FCFA)"
        ws2['C3'] = "Market Share"
        self._style_excel_row(ws2, 3, bold=True)
        
        row = 4
        for operadora, valor in receitas.get('por_operadora', {}).items():
            share = receitas.get('quota_receitas', {}).get(operadora, 0)
            ws2[f'A{row}'] = operadora
            ws2[f'B{row}'] = valor
            ws2[f'C{row}'] = f'{share:.1f}%'
            row += 1
        
        self._auto_size_columns(ws2)
    
    def _build_executive_report_excel(self, wb):
        """Constrói dashboard executivo em Excel"""
        ws = wb.create_sheet("Dashboard Executivo")
        self._style_excel_header(ws, f"Dashboard Executivo - {self.year}")
        
        kpis = self.report_data.get('kpis_principais', {})
        
        ws['A3'] = "Indicador"
        ws['B3'] = "Valor"
        ws['C3'] = "Variação"
        self._style_excel_row(ws, 3, bold=True)
        
        row = 4
        for key, kpi in kpis.items():
            nome = key.replace('_', ' ').title()
            ws[f'A{row}'] = nome
            ws[f'B{row}'] = str(kpi.get('valor', '--'))
            ws[f'C{row}'] = str(kpi.get('variacao', '--'))
            row += 1
        
        self._auto_size_columns(ws)
    
    def _build_comparative_report_excel(self, wb):
        """Constrói relatório comparativo em Excel"""
        ws = wb.create_sheet("Comparativo")
        self._style_excel_header(ws, f"Análise Comparativa - {self.year}")
        
        ws['A3'] = "Operadora"
        ws['B3'] = "Assinantes"
        ws['C3'] = "Market Share"
        ws['D3'] = "Tráfego Voz"
        self._style_excel_row(ws, 3, bold=True)
        
        if isinstance(self.report_data, list):
            row = 4
            for item in self.report_data:
                ws[f'A{row}'] = item.get('operadora', '--')
                ws[f'B{row}'] = item.get('assinantes', 0)
                ws[f'C{row}'] = item.get('market_share', '--')
                ws[f'D{row}'] = item.get('trafego_voz', 0)
                row += 1
        
        self._auto_size_columns(ws)
    
    # ===== CONSTRUÇÃO DE RELATÓRIOS CSV =====
    
    def _build_market_report_csv(self, writer):
        """Constrói relatório de mercado em CSV"""
        panorama = self.report_data.get('panorama_geral', {})
        
        writer.writerow(['Panorama Geral'])
        writer.writerow(['Total de Assinantes', panorama.get('total_assinantes', 0)])
        writer.writerow(['Crescimento Anual', f"{panorama.get('crescimento_percentual', 0):.2f}%"])
        writer.writerow(['Taxa de Penetração', f"{panorama.get('taxa_penetracao', 0):.2f}%"])
        writer.writerow([])
        
        writer.writerow(['Market Share'])
        for operadora, share in panorama.get('market_share', {}).items():
            writer.writerow([operadora, f'{share:.2f}%'])
        writer.writerow([])
        
        receitas = self.report_data.get('receitas', {})
        writer.writerow(['Receitas por Operadora'])
        writer.writerow(['Operadora', 'Receita Total', 'Market Share'])
        for operadora, valor in receitas.get('por_operadora', {}).items():
            share = receitas.get('quota_receitas', {}).get(operadora, 0)
            writer.writerow([operadora, valor, f'{share:.1f}%'])
    
    def _build_executive_report_csv(self, writer):
        """Constrói dashboard executivo em CSV"""
        kpis = self.report_data.get('kpis_principais', {})
        
        writer.writerow(['KPIs Principais'])
        writer.writerow(['Indicador', 'Valor', 'Variação'])
        
        for key, kpi in kpis.items():
            nome = key.replace('_', ' ').title()
            writer.writerow([nome, kpi.get('valor', '--'), kpi.get('variacao', '--')])
    
    def _build_comparative_report_csv(self, writer):
        """Constrói relatório comparativo em CSV"""
        writer.writerow(['Comparação entre Operadoras'])
        writer.writerow(['Operadora', 'Assinantes', 'Market Share', 'Tráfego Voz'])
        
        if isinstance(self.report_data, list):
            for item in self.report_data:
                writer.writerow([
                    item.get('operadora', '--'),
                    item.get('assinantes', 0),
                    item.get('market_share', '--'),
                    item.get('trafego_voz', 0)
                ])
    
    # ===== ESTILOS EXCEL =====
    
    def _style_excel_header(self, ws, title):
        """Aplica estilo ao cabeçalho do Excel"""
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True, color='F0B90B')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
    
    def _style_excel_row(self, ws, row, bold=False, bg_color=None):
        """Aplica estilo a uma linha do Excel"""
        for col in ['A', 'B', 'C', 'D']:
            cell = ws[f'{col}{row}']
            if bold:
                cell.font = Font(bold=True)
            if bg_color:
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    
    def _auto_size_columns(self, ws):
        """Ajusta largura das colunas automaticamente"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

